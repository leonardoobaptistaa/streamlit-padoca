# fmt:off
import streamlit as st # isort:skip
import pandas as pd
from conversor import dataframe_to_expense_report, dataframes_to_payment_report
from openpyxl import load_workbook
from service.payment_verifier import PaymentVerifier

st.set_page_config(
    page_title="Comparar gastos",
    page_icon="chart_with_upwards_trend",
) # isort:skip
# fmt:on


def create_payment_dataframe(file, sheet, wb):
    d1_value = wb[sheet]["D1"].value
    if d1_value == "PAGO":
        use_columns = [1, 2, 5]
    elif d1_value == "VENCIMENTO":
        use_columns = [1, 2, 7]

    dtype = {"Lançamento": str, "Valor": float, "Data": str}
    df = pd.read_excel(
        file,
        sheet_name=sheet,
        header=None,
        usecols=use_columns,
        names=["Lançamento", "Valor", "Data"],
        skiprows=1,
        dtype=dtype,
    )
    df.index = df.index + 1
    df.query("Data.notnull()", inplace=True)
    df.name = f"{file.name}/{sheet}"
    return df


def create_expense_dataframe(file):
    dtype = {"Data": str, "Lançamento": str, "Ag. origem": str, "Valor": float}
    df = pd.read_excel(
        file,
        sheet_name=0,
        header=None,
        usecols=[0, 1, 2, 3],
        names=["Data", "Lançamento", "Ag. origem", "Valor"],
        skiprows=10,
        dtype=dtype,
    )
    df.index = df.index + 11
    df.query("Valor < 0", inplace=True)
    df.name = f"{file.name}"
    return df


with st.form("my_form"):
    st.write("Escolha primeiro o extrato e depois um ou mais arquivos de despesas.")
    bank_report = st.file_uploader("Extrato")
    spends = st.file_uploader("Arquivo(s) de despesas", accept_multiple_files=True)

    submitted = st.form_submit_button("Comparar")
    if submitted:
        if not bank_report:
            st.warning("Está faltando o arquivo de extrato.", icon="⚠️")

        if not spends:
            st.warning("Está faltando arquivo(s) de despesas.", icon="⚠️")

        dfs = list()
        for spend_file in spends:
            wb = load_workbook(spend_file, read_only=True)
            selected_sheets = st.multiselect(spend_file.name, wb.sheetnames, [])

            use_columns = [0, 1, 2, 3]
            for selected_sheet in selected_sheets:
                df = create_payment_dataframe(spend_file, selected_sheet, wb)
                dfs.append(df)

        payment_report = dataframes_to_payment_report(dfs)

        df = create_expense_dataframe(bank_report)
        expense_report = dataframe_to_expense_report(df)

        payment_verifier = PaymentVerifier(expense_report, payment_report)
        payment_verifier.validate_payments()

        payments_expander = st.expander("Pagamentos", expanded=True)

        not_verified_only = payments_expander.checkbox(
            "Mostrar somente pagamentos não verificados.", value=True
        )

        df = pd.DataFrame(
            [x.to_present() for x in payment_verifier.payment_report.payments],
            columns=["Data pagamento", "Lançamento", "Total", "Validado", "Planilha"],
        )

        if not_verified_only:
            df.query("Validado == False", inplace=True)

        df = df.style.format(precision=2, thousands=".", decimal=",")
        payments_expander.dataframe(df, use_container_width=True, hide_index=True)

        debit_expander = st.expander("Débitos", expanded=True)
        not_verified_only = debit_expander.checkbox(
            "Mostrar somente débitos não verificados.", value=True
        )
        df = pd.DataFrame(
            [x.to_present() for x in payment_verifier.expense_report.expenses],
            columns=["Data", "Lançamento", "Total", "Validado", "Planilha"],
        )

        if not_verified_only:
            df.query("Validado == False", inplace=True)

        df = df.style.format(precision=2, thousands=".", decimal=",")
        debit_expander.dataframe(df, hide_index=True)
